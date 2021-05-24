import openpyxl
import csv
import pandas


class CSVModule(object):
    def import_csv(self, csv_path):
        with open(csv_path, 'r', newline='') as rds:
            list_ = []
            for rd in csv.reader(rds):
                try:
                    list_.append(rd)
                except Exception as ex:
                    print(ex)

            return list_


class XlsxModule(object):
    def __init__(self, path, new=False):
        self._col = 0
        self._row = 0
        self._path = path
        if new:
            self._wb = openpyxl.Workbook()
            self._ws = self._wb.active
            self._create_sheet()
        else:
            self._wb = openpyxl.load_workbook(path)
            self._ws = self._wb.active

        self._doc = openpyxl.load_workbook(path)

    def _create_sheet(self):
        self._wb.create_sheet('Sheet1')
        self._wb.save(self._path)

    def export_xlsx_to_row(self, data):
        try:
            if isinstance(data, list):
                for each in data:
                    self._ws.append(tuple(each))
                    self._wb.save(self._path)
            else:
                self._ws.append(tuple(data))
                self._wb.save(self._path)

        except Exception as ex:
            raise ex
